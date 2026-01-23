import os
import cv2
import numpy as np
from PIL import Image
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# === PARÁMETROS ===
mask_dir = r"D:\Python proyectos 2025\CNN EMANUEL\archive\generated_masks"
original_dir = r"D:\Python proyectos 2025\CNN EMANUEL\archive\train\Positive"
save_dir = r"D:\Python proyectos 2025\CNN EMANUEL\outputs_final_v2"
excel_path = os.path.join(save_dir, "diagnostico_grietas.xlsx")
os.makedirs(save_dir, exist_ok=True)

# === CONVERSIÓN DE UNIDADES (96 ppi) ===
pixels_per_cm = 96 / 2.54

# === INICIAR EXCEL ===
wb = Workbook()
ws = wb.active
ws.title = "Grietas Segmentadas"
ws.append(["Nombre", "Área (cm²)", "Mayor dimensión (cm)", "Ancho promedio (cm)", "Severidad", "Imagen resultante"])

# === FUNCIONES AUXILIARES ===
def insertar_imagen_excel(ws, img_path, row, col=6, max_width=120):
    img_excel = ExcelImage(img_path)
    img_excel.width = max_width
    img_excel.height = max_width
    cell = f"{get_column_letter(col)}{row}"
    ws.add_image(img_excel, cell)

def clasificar_severidad(ancho_cm):
    if ancho_cm < 0.15:
        return "Menor"
    elif 0.15 <= ancho_cm <= 0.3:
        return "Moderada"
    else:
        return "Severa"

# === PROCESAR CADA MÁSCARA ===
for filename in os.listdir(mask_dir):
    if filename.endswith(".png"):
        name = os.path.splitext(filename)[0]
        mask_path = os.path.join(mask_dir, filename)
        orig_path = os.path.join(original_dir, name + ".jpg")
        if not os.path.exists(orig_path):
            continue

        # Cargar imagen y máscara
        image = cv2.imread(orig_path)
        mask = cv2.imread(mask_path, cv2.IMREAD_GRAYSCALE)

        # Binarizar y limpiar
        _, binary_mask = cv2.threshold(mask, 127, 255, cv2.THRESH_BINARY)
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))
        cleaned_mask = cv2.morphologyEx(binary_mask, cv2.MORPH_OPEN, kernel)

        # Contornos y área
        contours, _ = cv2.findContours(cleaned_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        total_area_px = np.sum(cleaned_mask == 255)
        total_area_cm2 = total_area_px / (pixels_per_cm ** 2)

        max_length_cm = 0
        widths_cm = []

        overlay = image.copy()
        blue_mask = cv2.merge([cleaned_mask // 3, 0 * cleaned_mask, cleaned_mask])
        overlay = cv2.addWeighted(overlay, 1.0, blue_mask, 0.6, 0)

        for cnt in contours:
            if cv2.contourArea(cnt) < 10:
                continue

            rect = cv2.minAreaRect(cnt)
            box = cv2.boxPoints(rect)
            box = np.int32(box)

            width_px = min(rect[1])
            height_px = max(rect[1])
            width_cm = width_px / pixels_per_cm
            length_cm = height_px / pixels_per_cm

            widths_cm.append(width_cm)
            max_length_cm = max(max_length_cm, length_cm)

            # Dibujar caja y medidas
            cv2.drawContours(overlay, [box], 0, (0, 255, 0), 2)
            cv2.putText(overlay, f"{width_cm:.1f}cm", tuple(np.int32(box[0]) + [0, -5]),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 1)

        # Anotar área
        cv2.putText(overlay, f"Área total ~ {total_area_px} px", (10, overlay.shape[0] - 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)

        # Guardar imagen
        out_path = os.path.join(save_dir, f"{name}_result.png")
        cv2.imwrite(out_path, overlay)

        # === Agregar a Excel ===
        avg_width_cm = np.mean(widths_cm) if widths_cm else 0
        severidad = clasificar_severidad(avg_width_cm)
        ws.append([name, round(total_area_cm2, 2), round(max_length_cm, 2), round(avg_width_cm, 2), severidad])
        insertar_imagen_excel(ws, out_path, ws.max_row)

# === FORMATOS ===
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws.column_dimensions[col].width = 18

# === COLORES CONDICIONALES ===
for row in range(2, ws.max_row + 1):
    severidad = ws[f"E{row}"].value
    if severidad == "Severa":
        fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    elif severidad == "Moderada":
        fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    else:
        fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
    ws[f"E{row}"].fill = fill

# === GUARDAR EXCEL ===
wb.save(excel_path)
print(f"[✓] Proceso completo. Resultados guardados en:\n    {excel_path}")
