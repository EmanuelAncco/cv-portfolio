import tkinter as tk
from tkinter import simpledialog, messagebox, filedialog, ttk
from PIL import Image, ImageTk, ImageOps
import cv2
import os
import threading
import logging
import shutil
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

# --- Dependencias Clave ---
# Asegúrate de haberlas instalado con: pip install "numpy<2" opencv-python Pillow deepface openpyxl ultralytics
from deepface import DeepFace
from ultralytics import YOLO

# --- CONFIGURACIÓN DEL LOGGING ---
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] - %(message)s',
    handlers=[
        logging.FileHandler("access_control_v3.1.log", mode='w'),
        logging.StreamHandler()
    ]
)


class AccessControlApp(tk.Tk):
    """
    Una aplicación de escritorio completa que integra Reconocimiento Facial y
    Detección de EPP para el proyecto EMARC VISIÓN.
    """
    MANDATORY_EPP = {"Helmet"}

    def __init__(self):
        super().__init__()
        self.title("Central de Control de Acceso EMARC v3.1")
        self.geometry("1600x900")
        self.configure(bg="#2E2E2E")

        self.db_dir = "database"
        self.captures_dir = "verification_captures"
        os.makedirs(self.db_dir, exist_ok=True)
        os.makedirs(self.captures_dir, exist_ok=True)

        self.vid = cv2.VideoCapture(0)
        self.current_frame = None
        self.is_capturing = True

        self.attendance_log = {}
        self.logged_in_workers = set()

        self.epp_model = None
        self.load_epp_model()

        self.create_styles()
        self.create_widgets()

        self.update_webcam_feed()
        self.protocol("WM_DELETE_WINDOW", self.on_closing)

    def load_epp_model(self):
        """
        Carga el modelo de detección de EPP más reciente en un hilo separado.
        """

        def loader():
            logging.info("Buscando el modelo de EPP más reciente...")
            try:
                runs_dir = 'runs/detect'
                if not os.path.exists(runs_dir):
                    logging.error("Directorio 'runs/detect' no encontrado.")
                    return

                all_run_dirs = [d for d in os.listdir(runs_dir) if os.path.isdir(os.path.join(runs_dir, d))]
                if not all_run_dirs:
                    logging.error("No se encontraron ejecuciones de entrenamiento en 'runs/detect'.")
                    return

                latest_run_dir = max(all_run_dirs, key=lambda d: os.path.getmtime(os.path.join(runs_dir, d)))
                model_path = os.path.join(runs_dir, latest_run_dir, 'weights', 'best.pt')

                if os.path.exists(model_path):
                    logging.info(f"Cargando modelo de EPP desde: {model_path}")
                    self.epp_model = YOLO(model_path)
                    logging.info("¡Modelo de EPP cargado exitosamente!")
                else:
                    logging.error(f"No se encontró 'best.pt' en {latest_run_dir}")

            except Exception as e:
                logging.error(f"Error al cargar el modelo de EPP: {e}", exc_info=True)

        threading.Thread(target=loader, daemon=True).start()

    def create_styles(self):
        style = ttk.Style(self)
        style.theme_use('clam')
        self.accent_color = "#007ACC"
        self.bg_color = "#2E2E2E"
        self.text_color = "white"
        self.success_color = "#28a745"
        self.error_color = "#dc3545"
        self.warning_color = "#ffc107"

        style.configure("TFrame", background=self.bg_color)
        style.configure("TButton", padding=10, relief="flat", background="#5E5E5E", foreground=self.text_color,
                        font=('Helvetica', 12, 'bold'))
        style.map("TButton", background=[('active', '#7A7A7A')])
        style.configure("Accent.TButton", background=self.accent_color)
        style.map("Accent.TButton", background=[('active', '#005f9e')])
        style.configure("TLabel", background=self.bg_color, foreground=self.text_color, font=('Helvetica', 11))
        style.configure("Header.TLabel", font=('Helvetica', 18, 'bold'))
        style.configure("Result.TLabel", font=('Helvetica', 14, 'bold'))
        style.configure("Treeview", background="#3C3C3C", foreground=self.text_color, fieldbackground="#3C3C3C",
                        rowheight=25)
        style.map("Treeview", background=[('selected', self.accent_color)])
        style.configure("Treeview.Heading", font=('Helvetica', 12, 'bold'), background="#5E5E5E",
                        foreground=self.text_color)

    def create_widgets(self):
        main_frame = ttk.Frame(self, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        main_frame.grid_columnconfigure(0, weight=2)
        main_frame.grid_columnconfigure(1, weight=1)
        main_frame.grid_rowconfigure(1, weight=1)

        ttk.Label(main_frame, text="Central de Control de Acceso", style="Header.TLabel").grid(row=0, column=0,
                                                                                               columnspan=2,
                                                                                               pady=(0, 20))

        # --- Columna Izquierda: Webcam y Verificación ---
        left_column = ttk.Frame(main_frame)
        left_column.grid(row=1, column=0, sticky="nsew", padx=(0, 10))
        left_column.grid_rowconfigure(0, weight=3)

        webcam_main_frame = ttk.Frame(left_column)
        webcam_main_frame.grid(row=0, column=0, sticky="nsew")
        webcam_main_frame.grid_columnconfigure(0, weight=1)
        webcam_main_frame.grid_rowconfigure(0, weight=1)

        self.webcam_label = ttk.Label(webcam_main_frame)
        self.webcam_label.grid(row=0, column=0, sticky="nsew")

        # --- NUEVO (v3.1): Ventana de Detección de EPP en tiempo real ---
        self.epp_preview_label = ttk.Label(webcam_main_frame, relief="solid")
        # Posicionar en la esquina superior derecha
        self.epp_preview_label.place(relx=1.0, rely=0.0, x=-10, y=10, anchor="ne")

        control_frame = ttk.Frame(left_column, padding=10)
        control_frame.grid(row=1, column=0, sticky="ew", pady=(10, 0))
        control_frame.grid_columnconfigure((0, 1), weight=1)

        verify_file_button = ttk.Button(control_frame, text="Verificar (Archivo)", command=self.verify_from_file)
        verify_file_button.grid(row=0, column=0, sticky="ew", padx=5)
        verify_webcam_button = ttk.Button(control_frame, text="Verificar (Webcam)", command=self.verify_from_webcam)
        verify_webcam_button.grid(row=0, column=1, sticky="ew", padx=5)

        # --- Columna Derecha: Registro, Log y Exportación ---
        right_column = ttk.Frame(main_frame)
        right_column.grid(row=1, column=1, sticky="nsew", padx=(10, 0))
        right_column.grid_rowconfigure(2, weight=1)

        register_frame = ttk.Frame(right_column)
        register_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        register_frame.grid_columnconfigure((0, 1), weight=1)
        enroll_webcam_button = ttk.Button(register_frame, text="Registrar (Webcam)",
                                          command=self.start_enrollment_webcam, style="Accent.TButton")
        enroll_webcam_button.grid(row=0, column=0, sticky="ew", padx=5)
        enroll_file_button = ttk.Button(register_frame, text="Registrar (Archivos)",
                                        command=self.start_enrollment_files, style="Accent.TButton")
        enroll_file_button.grid(row=0, column=1, sticky="ew", padx=5)

        last_event_frame = ttk.Frame(right_column, padding=10, relief="solid")
        last_event_frame.grid(row=1, column=0, sticky="ew", pady=(10, 20))
        last_event_frame.grid_columnconfigure(1, weight=1)

        self.last_event_photo_label = ttk.Label(last_event_frame, background="#1C1C1C")
        self.last_event_photo_label.grid(row=0, column=0, rowspan=2, padx=10, pady=10)
        self.last_event_status_label = ttk.Label(last_event_frame, text="Esperando Verificación...",
                                                 font=('Helvetica', 16, 'bold'))
        self.last_event_status_label.grid(row=0, column=1, sticky="sw", padx=10)
        self.last_event_info_label = ttk.Label(last_event_frame, text="---")
        self.last_event_info_label.grid(row=1, column=1, sticky="nw", padx=10)

        log_frame = ttk.Frame(right_column)
        log_frame.grid(row=2, column=0, sticky="nsew")
        log_frame.grid_rowconfigure(1, weight=1)
        log_frame.grid_columnconfigure(0, weight=1)

        ttk.Label(log_frame, text="Historial de Asistencia del Día", font=('Helvetica', 14, 'bold')).grid(row=0,
                                                                                                          column=0,
                                                                                                          sticky="w",
                                                                                                          pady=(0, 5))

        cols = ("Nombre", "Hora de Ingreso", "Hora de Salida")
        self.log_tree = ttk.Treeview(log_frame, columns=cols, show="headings")
        for col in cols:
            self.log_tree.heading(col, text=col)
            self.log_tree.column(col, width=150, anchor="center")
        self.log_tree.grid(row=1, column=0, sticky="nsew")

        export_button = ttk.Button(right_column, text="Exportar a Excel", command=self.export_to_excel)
        export_button.grid(row=3, column=0, sticky="ew", pady=(10, 0))

    def update_webcam_feed(self):
        if not self.is_capturing: return
        ret, frame = self.vid.read()
        if ret:
            self.current_frame = frame

            # --- NUEVO (v3.1): Detección de EPP en tiempo real para la vista previa ---
            if self.epp_model:
                epp_results = self.epp_model.predict(source=frame, imgsz=320, verbose=False)
                annotated_frame = epp_results[0].plot()
                epp_preview_img = cv2.resize(annotated_frame, (213, 160))  # Tamaño pequeño para la esquina
                self.epp_photo = ImageTk.PhotoImage(
                    image=Image.fromarray(cv2.cvtColor(epp_preview_img, cv2.COLOR_BGR2RGB)))
                self.epp_preview_label.config(image=self.epp_photo)

            display_frame = cv2.resize(frame, (853, 640))
            self.photo = ImageTk.PhotoImage(image=Image.fromarray(cv2.cvtColor(display_frame, cv2.COLOR_BGR2RGB)))
            self.webcam_label.config(image=self.photo)

        self.after(30, self.update_webcam_feed)  # Aumentar ligeramente el delay para dar tiempo a la CPU

    def start_enrollment_webcam(self):
        worker_name = simpledialog.askstring("Registro (Webcam)",
                                             "Ingrese el nombre del trabajador (ej. Emanuel_Ancco):", parent=self)
        if not worker_name or not worker_name.strip(): return

        worker_dir = os.path.join(self.db_dir, worker_name)
        os.makedirs(worker_dir, exist_ok=True)

        for i in range(5):
            messagebox.showinfo("Captura de Registro", f"Se tomará la foto #{i + 1}/5. Centre su rostro y presione OK.")
            if self.current_frame is not None:
                cv2.imwrite(os.path.join(worker_dir, f"{i + 1}.jpg"), self.current_frame)

        messagebox.showinfo("Éxito", f"Se han registrado 5 imágenes para {worker_name}.")

    def start_enrollment_files(self):
        worker_name = simpledialog.askstring("Registro (Archivos)",
                                             "Ingrese el nombre del trabajador (ej. Emanuel_Ancco):", parent=self)
        if not worker_name or not worker_name.strip(): return

        files = filedialog.askopenfilenames(title="Seleccione las fotos del trabajador",
                                            filetypes=[("Archivos de Imagen", "*.jpg *.jpeg *.png")])
        if not files: return

        worker_dir = os.path.join(self.db_dir, worker_name)
        os.makedirs(worker_dir, exist_ok=True)

        for i, file_path in enumerate(files):
            try:
                img = Image.open(file_path)
                img = ImageOps.exif_transpose(img)
                img.convert("RGB").save(os.path.join(worker_dir, f"uploaded_{i + 1}.jpg"), "JPEG")
            except Exception as e:
                logging.error(f"No se pudo procesar el archivo {file_path}: {e}")

        messagebox.showinfo("Éxito", f"Se han registrado {len(files)} imágenes para {worker_name}.")

    def verify_from_file(self):
        image_path = filedialog.askopenfilename(title="Seleccione la imagen de prueba")
        if not image_path: return

        try:
            img = Image.open(image_path)
            img = ImageOps.exif_transpose(img)
            oriented_path = os.path.join(self.captures_dir, "temp_oriented_verify.jpg")
            img.convert("RGB").save(oriented_path, "JPEG")
            threading.Thread(target=self.run_verification, args=(oriented_path,)).start()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir la imagen: {e}")

    def verify_from_webcam(self):
        if self.current_frame is None: return
        temp_path = os.path.join(self.captures_dir, "temp_webcam_verify.jpg")
        cv2.imwrite(temp_path, self.current_frame)
        threading.Thread(target=self.run_verification, args=(temp_path,)).start()

    def run_verification(self, image_path):
        """
        Realiza el proceso de verificación de 2 pasos: Facial y EPP.
        """
        if self.epp_model is None:
            self.update_last_event_ui("X Error", "Modelo de EPP no cargado", self.error_color, image_path)
            return

        try:
            # --- PASO 1: Verificación Facial ---
            # --- CAMBIO (v3.1): enforce_detection=False para permitir rostros con casco ---
            dfs = DeepFace.find(img_path=image_path, db_path=self.db_dir, model_name='ArcFace',
                                distance_metric='cosine', enforce_detection=False)

            if not dfs or dfs[0].empty:
                self.update_last_event_ui("X Acceso Denegado", "NO RECONOCIDO", self.error_color, image_path)
                return

            result = dfs[0].iloc[0]
            identity_path = result['identity']
            distance = result['distance']
            worker_name = os.path.basename(os.path.dirname(identity_path))

            # --- CAMBIO (v3.1): Umbral ajustado a 0.6 ---
            # Nota: Un valor más alto es MENOS estricto para la distancia de coseno.
            threshold = 0.6
            if distance > threshold:
                info_text = f"Coincidencia: {worker_name} ({distance:.2f})"
                self.update_last_event_ui("X Acceso Denegado", info_text, self.warning_color, image_path)
                return

            # --- PASO 2: Auditoría de EPP (si la identidad es correcta) ---
            logging.info(f"Identidad verificada para {worker_name}. Realizando auditoría de EPP...")

            epp_results = self.epp_model.predict(source=image_path, imgsz=640, verbose=False)
            detected_classes = set()
            if epp_results and epp_results[0].boxes:
                class_ids = epp_results[0].boxes.cls.cpu().numpy().astype(int)
                for cid in class_ids:
                    detected_classes.add(self.epp_model.names[cid])

            logging.info(f"EPPs detectados: {detected_classes}")

            missing_epp = self.MANDATORY_EPP - detected_classes

            if not missing_epp:
                self.log_attendance(worker_name, image_path)
            else:
                missing_str = ", ".join(missing_epp)
                info_text = f"{worker_name}: Falta {missing_str}"
                self.update_last_event_ui("X EPP Incompleto", info_text, self.error_color, image_path)

        except ValueError as e:
            logging.error(f"Error de DeepFace: {e}")
            self.update_last_event_ui("X Error", "No se detectó rostro", self.error_color, image_path)
        except Exception as e:
            logging.error(f"Error inesperado en verificación: {e}", exc_info=True)
            self.update_last_event_ui("X Error", "Error en el proceso", self.error_color, image_path)

    def log_attendance(self, worker_name, image_path):
        """
        Gestiona el registro de ingresos y salidas.
        """
        now = datetime.now()
        today = now.strftime("%Y-%m-%d")
        current_time_str = now.strftime("%H:%M:%S")

        if worker_name not in self.attendance_log: self.attendance_log[worker_name] = {}
        if today not in self.attendance_log[worker_name]: self.attendance_log[worker_name][today] = {}

        log_entry = self.attendance_log[worker_name][today]

        timestamp = now.strftime("%Y%m%d_%H%M%S")
        event_type = "salida" if worker_name in self.logged_in_workers else "ingreso"
        unique_filename = f"{worker_name}_{timestamp}_{event_type}.jpg"
        persistent_photo_path = os.path.join(self.captures_dir, unique_filename)
        shutil.copy(image_path, persistent_photo_path)

        if worker_name not in self.logged_in_workers:
            log_entry['entry_time'] = current_time_str
            log_entry['entry_photo_path'] = persistent_photo_path
            self.logged_in_workers.add(worker_name)
            status_text = f"✓ Ingreso Registrado"
            info_text = f"{worker_name} a las {current_time_str}"
            self.update_last_event_ui(status_text, info_text, self.success_color, persistent_photo_path)
        else:
            log_entry['exit_time'] = current_time_str
            log_entry['exit_photo_path'] = persistent_photo_path
            self.logged_in_workers.remove(worker_name)
            status_text = f"✓ Salida Registrada"
            info_text = f"{worker_name} a las {current_time_str}"
            self.update_last_event_ui(status_text, info_text, self.success_color, persistent_photo_path)

        self.update_log_tree()

    def update_last_event_ui(self, status, info, color, photo_path):
        """
        Actualiza el panel de "Último Evento" con la información de la verificación.
        """
        self.last_event_status_label.config(text=status, foreground=color)
        self.last_event_info_label.config(text=info, foreground=self.text_color)

        try:
            img = Image.open(photo_path)
            img.thumbnail((100, 100))
            photo = ImageTk.PhotoImage(img)
            self.last_event_photo_label.config(image=photo)
            self.last_event_photo_label.image = photo
        except Exception as e:
            logging.error(f"No se pudo mostrar la foto del último evento: {e}")

    def update_log_tree(self):
        """
        Actualiza la tabla de asistencia en la UI.
        """
        for i in self.log_tree.get_children():
            self.log_tree.delete(i)

        today = datetime.now().strftime("%Y-%m-%d")

        sorted_logs = sorted(
            self.attendance_log.items(),
            key=lambda item: item[1].get(today, {}).get('entry_time', '23:59:59')
        )

        for worker_name, dates in sorted_logs:
            if today in dates:
                log = dates[today]
                entry_time = log.get('entry_time', '---')
                exit_time = log.get('exit_time', '---')
                self.log_tree.insert("", "end", values=(worker_name, entry_time, exit_time))

    def export_to_excel(self):
        """
        NUEVO (v3.1): Exporta el registro de asistencia a un archivo Excel con formato profesional.
        """
        if not self.attendance_log:
            messagebox.showinfo("Vacío", "No hay registros de asistencia para exportar.")
            return

        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Workbook", "*.xlsx")],
                                                 title="Guardar Reporte de Asistencia")
        if not save_path: return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Registro Detallado"

            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            center_alignment = Alignment(horizontal='center', vertical='center')

            # --- Hoja de Registro Detallado ---
            headers = ["Trabajador", "Fecha", "Hora de Ingreso", "Foto Ingreso", "Hora de Salida", "Foto Salida"]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment

            for i, _ in enumerate(headers):
                ws.column_dimensions[get_column_letter(i + 1)].width = 22

            row_num = 2
            for worker_name, dates in self.attendance_log.items():
                for date, log in dates.items():
                    ws.row_dimensions[row_num].height = 95

                    ws.cell(row=row_num, column=1, value=worker_name).alignment = center_alignment
                    ws.cell(row=row_num, column=2, value=date).alignment = center_alignment
                    ws.cell(row=row_num, column=3, value=log.get('entry_time', '')).alignment = center_alignment
                    ws.cell(row=row_num, column=5, value=log.get('exit_time', '')).alignment = center_alignment

                    if 'entry_photo_path' in log and os.path.exists(log['entry_photo_path']):
                        img = OpenpyxlImage(log['entry_photo_path'])
                        img.height = 120
                        img.width = 120
                        ws.add_image(img, f'D{row_num}')

                    if 'exit_photo_path' in log and os.path.exists(log['exit_photo_path']):
                        img = OpenpyxlImage(log['exit_photo_path'])
                        img.height = 120
                        img.width = 120
                        ws.add_image(img, f'F{row_num}')

                    row_num += 1

            # --- Hoja de Resumen ---
            summary_ws = wb.create_sheet(title="Resumen Semanal")
            summary_headers = ["Fecha", "Trabajador", "Horas Trabajadas (aprox)"]
            summary_ws.append(summary_headers)
            for cell in summary_ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            for i, _ in enumerate(summary_headers):
                summary_ws.column_dimensions[get_column_letter(i + 1)].width = 30

            summary_row = 2
            for worker_name, dates in self.attendance_log.items():
                for date, log in dates.items():
                    if 'entry_time' in log and 'exit_time' in log:
                        try:
                            t1 = datetime.strptime(log['entry_time'], "%H:%M:%S")
                            t2 = datetime.strptime(log['exit_time'], "%H:%M:%S")
                            hours_worked = (t2 - t1)
                        except:
                            hours_worked = "Error en cálculo"

                        summary_ws.cell(row=summary_row, column=1, value=date)
                        summary_ws.cell(row=summary_row, column=2, value=worker_name)
                        summary_ws.cell(row=summary_row, column=3, value=str(hours_worked))
                        summary_row += 1

            wb.save(save_path)
            messagebox.showinfo("Éxito", f"Reporte de asistencia guardado en:\n{save_path}")
        except Exception as e:
            logging.error(f"Error al exportar a Excel: {e}", exc_info=True)
            messagebox.showerror("Error de Exportación", f"No se pudo guardar el archivo:\n{e}")

    def on_closing(self):
        self.is_capturing = False
        if self.vid.isOpened():
            self.vid.release()
        self.destroy()


if __name__ == "__main__":
    app = AccessControlApp()
    app.mainloop()
