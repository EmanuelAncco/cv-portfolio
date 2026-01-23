import os
import cv2
import numpy as np
from PIL import Image
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import math

# -----------------------------
# Parámetros de configuración
# -----------------------------
dpi = 96  # asumiendo 96 ppp
cm_per_pixel = 2.54 / dpi
mask_dir = r"D:\Python proyectos 2025\CNN EMANUEL\archive\generated_masks"
image_dir = r"D:\Python proyectos 2025\CNN EMANUEL\archive\train\Positive"
save_dir = r"D:\Python proyectos 2025\CNN EMANUEL\outputs_final_v3"
excel_path = os.path.join(save_dir, "diagnostico_grietas.xlsx")

os.makedirs(save_dir, exist_ok=True)

# Para clasificar severidad
def clasificar_severidad(ancho_cm):
    if ancho_cm < 0.15:
        return "Menor"
    elif ancho_cm < 0.3:
        return "Moderada"
    else:
        return "Severa"

# -----------------------------
# Crear Excel
# -----------------------------
wb = Workbook()
ws = wb.active
ws.title = "Grietas Segmentadas"

# Encabezados
headers = ["Nombre", "Área (cm²)", "Mayor dimensión (cm)", "Ancho promedio (cm)", "Daño (%)", "Severidad", "Imagen resultante"]
ws.append(headers)

# Columnas con nombres fijos
for col, h in enumerate(headers, 1):
    ws.cell(1, col, h)

# -----------------------------
# Procesar
# -----------------------------
files = [f for f in os.listdir(mask_dir) if f.endswith(".png")]
files.sort()

for file in files:
    name = os.path.splitext(file)[0]
    mask_path = os.path.join(mask_dir, file)
    img_path = os.path.join(image_dir, f"{name}.jpg")

    if not os.path.exists(img_path):
        continue

    # Cargar imagen y máscara
    image = cv2.imread(img_path)
    if image is None:
        continue

    mask = cv2.imread(mask_path, cv2.IMREAD_GRAYSCALE)
    if mask is None:
        continue

    h_img, w_img = image.shape[:2]

    _, binary = cv2.threshold(mask, 127, 255, cv2.THRESH_BINARY)
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3,3))
    cleaned = cv2.morphologyEx(binary, cv2.MORPH_OPEN, kernel)

    # Área en px
    area_px = np.sum(cleaned == 255)
    # Área total imagen
    total_px = w_img * h_img
    # Área en cm2
    area_cm2 = area_px * (cm_per_pixel**2)
    # Porcentaje de daño
    damage_percent = (area_px / total_px) * 100

    contours, _ = cv2.findContours(cleaned, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    max_len_cm = 0
    widths = []

    # Generar overlay
    overlay = image.copy()
    # Pintar en color rosado la grieta
    pink = np.zeros_like(image, dtype=np.uint8)
    pink[cleaned==255] = (180, 0, 180)  # BGR (magenta)
    alpha = 0.5
    overlay = cv2.addWeighted(overlay, 1.0, pink, alpha, 0)

    for cnt in contours:
        area_cnt = cv2.contourArea(cnt)
        if area_cnt < 10:
            continue

        # Caja rotada
        rect = cv2.minAreaRect(cnt)
        box = cv2.boxPoints(rect).astype(int)
        (w, h) = rect[1]
        length = max(w, h) * cm_per_pixel
        width  = min(w, h) * cm_per_pixel
        if length > max_len_cm:
            max_len_cm = length
        widths.append(width)

        # Dibujar bounding box
        cv2.drawContours(overlay, [box], 0, (0, 255, 0), 2)
        # Escribir en la esquina un valor
        cv2.putText(overlay,
                    f"{width:.2f}cm",
                    tuple(box[0]),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.7, (0,255,0), 2, cv2.LINE_AA)

    avg_width = np.mean(widths) if len(widths)>0 else 0
    severidad = clasificar_severidad(avg_width)

    # Texto final en la parte baja
    # color amarillo (B,G,R => 0,255,255)
    text_color = (0,255,255)
    y_off = overlay.shape[0] - 70
    cv2.putText(overlay, f"Area: {area_cm2:.2f} cm2", (10, y_off),
                cv2.FONT_HERSHEY_SIMPLEX, 0.9, text_color, 2)
    cv2.putText(overlay, f"Largo: {max_len_cm:.2f} cm", (10, y_off+25),
                cv2.FONT_HERSHEY_SIMPLEX, 0.9, text_color, 2)
    cv2.putText(overlay, f"Ancho: {avg_width:.2f} cm", (10, y_off+50),
                cv2.FONT_HERSHEY_SIMPLEX, 0.9, text_color, 2)

    # Daño
    damage_text = f"Daño: {damage_percent:.1f}%"
    cv2.putText(overlay, damage_text, (250, y_off),
                cv2.FONT_HERSHEY_SIMPLEX, 0.9, text_color, 2)

    # Guardar imagen final
    out_path = os.path.join(save_dir, f"{name}_annotated.png")
    cv2.imwrite(out_path, overlay)

    # Añadir fila en Excel
    row = ws.max_row+1
    ws.cell(row, 1).value = name
    ws.cell(row, 2).value = round(area_cm2, 2)
    ws.cell(row, 3).value = round(max_len_cm,2)
    ws.cell(row, 4).value = round(avg_width,2)
    ws.cell(row, 5).value = round(damage_percent,2)
    ws.cell(row, 6).value = severidad

    # Insertar imagen en Excel
    from openpyxl.drawing.image import Image as XLImage
    img_excel = XLImage(out_path)
    # Ajustar escalas
    img_excel.width = 160
    img_excel.height = 160
    col_img = 7  # G
    cell_loc = f"{get_column_letter(col_img)}{row}"
    ws.add_image(img_excel, cell_loc)

# Ajustar ancho de columnas
col_widths = [12, 15, 15, 15, 15, 12, 25]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Ajustar altura de filas para las imágenes
for row_i in range(2, ws.max_row + 1):
    ws.row_dimensions[row_i].height = 120

# Resaltar celdas de severidad
for row_i in range(2, ws.max_row+1):
    severity_val = ws.cell(row_i, 6).value
    if severity_val == "Severa":
        fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        ws.cell(row_i, 6).fill = fill
    elif severity_val == "Moderada":
        fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        ws.cell(row_i, 6).fill = fill
    else:
        fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
        ws.cell(row_i, 6).fill = fill

# Guardar Excel
wb.save(excel_path)
print(f"[✅] Listo. Revisa tu carpeta:\n{save_dir}\nArchivo Excel:\n{excel_path}")
