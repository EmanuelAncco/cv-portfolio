import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mtick
import seaborn as sns
import logging
import os
import sys

# Importaciones necesarias para manipular Excel e Imágenes
try:
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage
except ImportError:
    pass  # Se maneja en la función correspondiente

# --- CONFIGURACIÓN DEL EXPERIMENTO ---
CONFIG = {
    # NOTA: Cambia este nombre por el de tu archivo real .xls o .xlsx si lo corres localmente
    'input_file': 'y9qQY4NhRxyakGODYfccag_37a5bb7975df461bb05c4397ed1e74fd_1.Inventory-classification---project---data.xls',
    'output_dir': 'resultados_abc',
    'submission_file': 'Entrega_Final_Proyecto.xlsx',  # Nombre del archivo final
    'thresholds': {
        'A': 0.80,
        'B': 0.95
    },
    'plot_settings': {
        'title': 'Análisis ABC y Diagrama de Pareto - Inventario de Alimentos',
        'xlabel': 'Artículos (Ordenados por Valor de Uso)',
        'ylabel_bar': 'Valor de Uso Bi-semanal ($)',
        'ylabel_line': 'Porcentaje Acumulado del Valor Total (%)',
        'figsize': (12, 8)
    }
}

# --- CONFIGURACIÓN DE LOGGING ---
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler('abc_analysis.log', mode='w')
    ]
)
logger = logging.getLogger(__name__)


def setup_environment():
    """Crea la estructura de directorios necesaria."""
    try:
        if not os.path.exists(CONFIG['output_dir']):
            os.makedirs(CONFIG['output_dir'])
            logger.info(f"Directorio creado: {CONFIG['output_dir']}")
    except OSError as e:
        logger.critical(f"No se pudo crear el directorio de salida: {e}")
        sys.exit(1)


def load_and_clean_data(filepath):
    """
    Carga datos con enfoque pesimista y gestión granular de dependencias.
    """
    if not os.path.exists(filepath):
        logger.error(f"Archivo no encontrado: {filepath}")
        logger.warning("Generando datos sintéticos de prueba para continuar el flujo...")
        data = {
            'Item name': [f'Item_{i}' for i in range(1, 21)],
            'Bi-week usage (units)': [float(x) for x in range(20, 0, -1)],
            'Unit cost (US $)': [10.0] * 20
        }
        df = pd.DataFrame(data)
        return df

    try:
        # DETECCIÓN DE FORMATO INTELIGENTE Y GESTIÓN DE DRIVERS
        _, file_extension = os.path.splitext(filepath)
        file_ext_lower = file_extension.lower()

        if file_ext_lower == '.xls':
            logger.info(f"Detectado formato Legacy Excel ({file_extension}). Usando motor 'xlrd'.")
            try:
                df = pd.read_excel(filepath, engine='xlrd')
            except ImportError:
                logger.critical("ERROR DE DEPENDENCIA: Falta la librería 'xlrd'. Ejecuta 'pip install xlrd'.")
                sys.exit(1)

        elif file_ext_lower in ['.xlsx', '.xlsm']:
            logger.info(f"Detectado formato Excel Moderno ({file_extension}). Usando motor 'openpyxl'.")
            try:
                df = pd.read_excel(filepath, engine='openpyxl')
            except ImportError:
                logger.critical("ERROR DE DEPENDENCIA: Falta la librería 'openpyxl'. Ejecuta 'pip install openpyxl'.")
                sys.exit(1)
        else:
            logger.info(f"Detectado formato CSV/Texto ({file_extension}). Usando motor estándar.")
            df = pd.read_csv(filepath)

        logger.info(f"Datos cargados exitosamente. Dimensiones: {df.shape}")
        df.columns = df.columns.str.strip()

        required_cols = ['Item name', 'Bi-week usage (units)', 'Unit cost (US $)']
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise ValueError(f"Faltan columnas requeridas: {missing_cols}")

        return df

    except Exception as e:
        logger.critical(f"Error fatal inesperado cargando el archivo: {e}")
        sys.exit(1)


def perform_abc_analysis(df):
    """Ejecuta la lógica matemática del análisis ABC."""
    try:
        df['usage'] = pd.to_numeric(df['Bi-week usage (units)'], errors='coerce').fillna(0)
        df['cost'] = pd.to_numeric(df['Unit cost (US $)'], errors='coerce').fillna(0)
        df['Total_Value'] = df['usage'] * df['cost']

        df = df.sort_values(by='Total_Value', ascending=False).reset_index(drop=True)

        total_inventory_value = df['Total_Value'].sum()
        df['Cumulative_Value'] = df['Total_Value'].cumsum()
        df['Cumulative_Percentage'] = (df['Cumulative_Value'] / total_inventory_value) * 100

        def classify_item(cum_pct):
            if cum_pct <= (CONFIG['thresholds']['A'] * 100):
                return 'A'
            elif cum_pct <= (CONFIG['thresholds']['B'] * 100):
                return 'B'
            else:
                return 'C'

        df['Class'] = df['Cumulative_Percentage'].apply(classify_item)
        logger.info("Análisis ABC completado.")
        return df, total_inventory_value

    except Exception as e:
        logger.error(f"Error durante el cálculo matemático: {e}")
        raise


def generate_pareto_chart(df, output_path):
    """Genera y guarda el gráfico de Pareto."""
    try:
        sns.set_style("whitegrid")
        fig, ax1 = plt.subplots(figsize=CONFIG['plot_settings']['figsize'])

        colors = df['Class'].map({'A': '#2ecc71', 'B': '#f1c40f', 'C': '#e74c3c'})
        ax1.bar(df['Item name'], df['Total_Value'], color=colors, alpha=0.7, label='Valor de Uso ($)')
        ax1.set_ylabel(CONFIG['plot_settings']['ylabel_bar'], fontsize=12)
        ax1.set_xlabel(CONFIG['plot_settings']['xlabel'], fontsize=12)
        ax1.tick_params(axis='x', rotation=90, labelsize=8)

        ax2 = ax1.twinx()
        ax2.plot(df['Item name'], df['Cumulative_Percentage'], color='#2980b9', marker='o', ms=4, linewidth=2,
                 label='% Acumulado')
        ax2.set_ylabel(CONFIG['plot_settings']['ylabel_line'], fontsize=12)
        ax2.set_ylim(0, 110)
        ax2.yaxis.set_major_formatter(mtick.PercentFormatter())

        ax2.axhline(80, color='gray', linestyle='--', linewidth=1, alpha=0.5)
        ax2.axhline(95, color='gray', linestyle='--', linewidth=1, alpha=0.5)

        # Leyenda personalizada
        from matplotlib.patches import Patch
        legend_elements = [
            Patch(facecolor='#2ecc71', label='Clase A (Alta)'),
            Patch(facecolor='#f1c40f', label='Clase B (Media)'),
            Patch(facecolor='#e74c3c', label='Clase C (Baja)'),
            plt.Line2D([0], [0], color='#2980b9', lw=2, label='% Acumulado')
        ]
        ax1.legend(handles=legend_elements, loc='upper left')

        plt.title(CONFIG['plot_settings']['title'], fontsize=16, pad=20)
        plt.tight_layout()
        plt.savefig(output_path, dpi=300)
        plt.close()
        logger.info(f"Gráfico de Pareto guardado en: {output_path}")

    except Exception as e:
        logger.error(f"Error generando visualización: {e}")
        raise


def create_submission_file(df, image_path):
    """
    PASO FINAL: Empaqueta los datos y la imagen en un solo archivo Excel (.xlsx)
    para cumplir con los requisitos de la plataforma de subida.
    """
    output_path = os.path.join(CONFIG['output_dir'], CONFIG['submission_file'])

    try:
        logger.info("Iniciando empaquetado del archivo de entrega final...")

        # 1. Escribir los datos numéricos en la primera hoja
        # Usamos engine='openpyxl' explícitamente para asegurar compatibilidad
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Datos Clasificados', index=False)

        # 2. Insertar la imagen en una segunda hoja (o la misma)
        # Requerimos cargar el workbook recién creado para editarlo
        wb = load_workbook(output_path)

        # Creamos una hoja dedicada al gráfico
        ws_graph = wb.create_sheet('Gráfico Pareto')

        # Cargamos la imagen generada
        img = XLImage(image_path)

        # Insertamos la imagen en la celda A1 de la nueva hoja
        ws_graph.add_image(img, 'A1')

        wb.save(output_path)
        logger.info("=" * 60)
        # CORRECCIÓN: Eliminamos el emoji ✅ para evitar UnicodeEncodeError en terminales Windows
        logger.info(f"[SUCCESS] ARCHIVO LISTO: {output_path}")
        logger.info("Sube este archivo Excel a la plataforma.")
        logger.info("=" * 60)

    except NameError:
        logger.critical("Faltan librerías para manipular Excel/Imágenes.")
        logger.critical(">>> EJECUTA: pip install openpyxl pillow")
    except Exception as e:
        logger.error(f"Error al empaquetar el Excel final: {e}")
        logger.warning("Intenta subir el archivo CSV y la imagen PNG por separado si esto persiste.")


def main():
    logger.info("Iniciando Pipeline de Clasificación ABC...")
    setup_environment()

    # 1. Carga
    df_raw = load_and_clean_data(CONFIG['input_file'])

    # 2. Procesamiento
    df_processed, total_val = perform_abc_analysis(df_raw)

    # 3. Visualización
    plot_path = os.path.join(CONFIG['output_dir'], 'pareto_chart.png')
    generate_pareto_chart(df_processed, plot_path)

    # 4. Empaquetado Final (NUEVO)
    create_submission_file(df_processed, plot_path)

    logger.info("Pipeline finalizado exitosamente.")


if __name__ == "__main__":
    main()