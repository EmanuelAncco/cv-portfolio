# generador_excel_valorizaciones.py
# Este script genera un archivo Excel para simular el cálculo de una valorización de obra en Perú.
# Utiliza la librería openpyxl para crear y dar formato al archivo.

import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def apply_header_style(cell):
    """Aplica un estilo de cabecera a una celda (fondo azul, texto blanco en negrita)."""
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))
    cell.border = thin_border


def apply_data_style(cell, is_currency=False, is_percentage=False):
    """Aplica un estilo de datos a una celda (bordes y formato de número)."""
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))
    cell.border = thin_border
    if is_currency:
        cell.number_format = '"S/" #,##0.00'
    if is_percentage:
        cell.number_format = '0.00%'


def auto_adjust_columns(worksheet):
    """Ajusta el ancho de las columnas de una hoja de cálculo al contenido."""
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Corrección: obtener la letra de la columna
        for cell in col:
            try:
                if cell.value:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width


# --- Creación del Libro de Trabajo ---
wb = openpyxl.Workbook()

# --- Hoja 4: Datos_Base ---
ws_datos = wb.active
ws_datos.title = "Datos_Base"

# Título de la hoja
title_cell = ws_datos['A1']
title_cell.value = "DATOS GENERALES DEL PROYECTO Y PARÁMETROS"
title_cell.font = Font(bold=True, size=14)

# Tabla de Datos del Contrato
header_cell = ws_datos['B4']
header_cell.value = "Datos del Contrato"
apply_header_style(header_cell)
ws_datos.merge_cells('B4:C4')

contract_data = {
    "Monto del Contrato (S/IGV):": 1000000.00,
    "Plazo de Ejecución (días):": 120,
    "% Adelanto Directo:": 0.10,
    "% IGV:": 0.18
}
row_start = 5
for label, value in contract_data.items():
    cell_label = ws_datos[f'B{row_start}']
    cell_label.value = label
    cell_value = ws_datos[f'C{row_start}']
    cell_value.value = value
    apply_data_style(cell_label)
    apply_data_style(cell_value, is_currency="Monto" in label, is_percentage="%" in label)
    row_start += 1

# Tabla de Índices Unificados de Precios (INEI)
header_cell_iu = ws_datos['B10']
header_cell_iu.value = "Índices Unificados de Precios (INEI) - Área Geográfica 2"
apply_header_style(header_cell_iu)
ws_datos.merge_cells('B10:E10')

iu_headers = ["Código IU", "Monomio", "Base (mes 0)", "Reajuste (mes 1)"]
iu_data = [
    (47, "Mano de Obra", 450.5, 455.2),
    (39, "Gastos Generales", 430.1, 433.5),
    (49, "Maquinaria", 380.9, 385.1)
]
ws_datos.append(iu_headers)
for i, header in enumerate(iu_headers, start=1):
    apply_header_style(ws_datos.cell(row=11, column=i+1)) # Ajuste de columna

for row_data in iu_data:
    ws_datos.append(("",) + row_data) # Añadir una tupla vacía para alinear con la columna B
for row in ws_datos['B12:E14']:
    for cell in row:
        apply_data_style(cell)

# --- Hoja 3: Calculo_Reajuste_K ---
ws_k = wb.create_sheet("Calculo_Reajuste_K")

title_cell_k = ws_k['A1']
title_cell_k.value = "CÁLCULO DEL COEFICIENTE DE REAJUSTE 'K'"
title_cell_k.font = Font(bold=True, size=14)

ws_k['A4'] = "Fórmula Polinómica de Ejemplo:"
ws_k['C4'] = "K = 0.400(MOr/MOo) + 0.350(GGr/GGo) + 0.250(MAr/MAo)"

k_headers = ["Componente", "Coeficiente (%)", "Índice IU", "Base (o)", "Reajuste (r)", "Relación (r/o)"]
ws_k.append(k_headers)
for i, header in enumerate(k_headers, start=1):
    apply_header_style(ws_k.cell(row=6, column=i))

k_data = [
    ("Mano de Obra", 0.400, 47),
    ("Gastos Generales", 0.350, 39),
    ("Maquinaria", 0.250, 49)
]
for i, data_row in enumerate(k_data, start=7):
    ws_k[f'A{i}'] = data_row[0]
    ws_k[f'B{i}'] = data_row[1]
    ws_k[f'C{i}'] = data_row[2]
    # Fórmulas para buscar los índices en la hoja Datos_Base
    ws_k[f'D{i}'] = f"=VLOOKUP(C{i},Datos_Base!$C$12:$E$14,2,FALSE)"
    ws_k[f'E{i}'] = f"=VLOOKUP(C{i},Datos_Base!$C$12:$E$14,3,FALSE)"
    # Fórmula para la relación r/o
    ws_k[f'F{i}'] = f"=E{i}/D{i}"

# Fórmula para calcular K
ws_k['E11'] = "K Calculado:"
ws_k['E11'].font = Font(bold=True)
ws_k['F11'] = "=SUMPRODUCT(B7:B9, F7:F9)"
ws_k['F11'].font = Font(bold=True, color="FF0000")
ws_k['F11'].number_format = '0.000'

for row in ws_k['A7:F9']:
    for cell in row:
        apply_data_style(cell)

# --- Hoja 2: Cuerpo_Valorizacion_01 ---
ws_cuerpo = wb.create_sheet("Cuerpo_Valorizacion_01")
ws_cuerpo['A1'] = "CUERPO DE LA VALORIZACIÓN N° 01"
ws_cuerpo['A1'].font = Font(bold=True, size=14)

cuerpo_headers = ["Item", "Descripción", "Und.", "Metrado", "Precio Unit.", "Parcial", "Metrado Anterior", "Metrado Actual", "Metrado Acumulado", "Saldo Metrado", "Valorizado Actual"]
ws_cuerpo.append(cuerpo_headers)
for i, h in enumerate(cuerpo_headers, 1):
    apply_header_style(ws_cuerpo.cell(row=3, column=i))

partidas = [
    "01 ESTRUCTURAS",
    ("01.01", "MOVIMIENTO DE TIERRAS", "", "", "", "", "", "", "", "", ""),
    ("01.01.01", "Excavación masiva", "m3", 100, 25.50, "", 0, 50, "", "", ""),
    ("01.02", "CONCRETO SIMPLE", "", "", "", "", "", "", "", "", ""),
    ("01.02.01", "Cimientos corridos", "m3", 60, 150.00, "", 0, 30, "", "", ""),
]
row_idx = 4
for p in partidas:
    if isinstance(p, str):  # Es un título
        cell = ws_cuerpo.cell(row=row_idx, column=1)
        cell.value = p
        cell.font = Font(bold=True)
        ws_cuerpo.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(cuerpo_headers))
    else:  # Es una partida con datos
        ws_cuerpo.append(p)
        # Fórmulas de Excel
        ws_cuerpo[f'F{row_idx}'] = f"=D{row_idx}*E{row_idx}"
        ws_cuerpo[f'I{row_idx}'] = f"=G{row_idx}+H{row_idx}"
        ws_cuerpo[f'J{row_idx}'] = f"=D{row_idx}-I{row_idx}"
        ws_cuerpo[f'K{row_idx}'] = f"=H{row_idx}*E{row_idx}"

        # Aplicar estilos
        for col_letter in ['E', 'F', 'K']:
            apply_data_style(ws_cuerpo[f'{col_letter}{row_idx}'], is_currency=True)
        for col_letter in ['A', 'B', 'C', 'D', 'G', 'H', 'I', 'J']:
            apply_data_style(ws_cuerpo[f'{col_letter}{row_idx}'])
    row_idx += 1

# Fila de Totales
total_row = row_idx
ws_cuerpo[f'J{total_row}'] = "TOTAL VALORIZADO S/"
ws_cuerpo[f'J{total_row}'].font = Font(bold=True)
ws_cuerpo[f'K{total_row}'] = f"=SUM(K4:K{total_row - 1})"
ws_cuerpo[f'K{total_row}'].font = Font(bold=True)
apply_data_style(ws_cuerpo[f'K{total_row}'], is_currency=True)

# --- Hoja 1: Resumen_Valorizacion_01 ---
ws_resumen = wb.create_sheet("Resumen_Valorizacion_01", 0) # Crear y mover al principio
title_cell_resumen = ws_resumen['A1']
title_cell_resumen.value = "RESUMEN DE VALORIZACIÓN N° 01"
title_cell_resumen.font = Font(bold=True, size=16)

resumen_structure = [
    ("MONTO CONTRATO", f"=Datos_Base!C5"),
    ("VALORIZACIÓN BRUTA", f"=Cuerpo_Valorizacion_01!K{total_row}"),
    ("REAJUSTE (K-1)", f"=(Calculo_Reajuste_K!F11-1)*B6"),
    ("SUBTOTAL", "=SUM(B6:B7)"),
    ("AMORTIZACIÓN ADELANTO DIRECTO", f"=-Datos_Base!C7*B6"),
    ("SUBTOTAL 2", "=SUM(B8:B9)"),
    ("IGV (18%)", f"=B10*Datos_Base!C8"),
    ("TOTAL A PAGAR", "=B10+B11"),
]
row_idx = 5
for label, formula in resumen_structure:
    cell_label = ws_resumen[f'A{row_idx}']
    cell_label.value = label
    cell_value = ws_resumen[f'B{row_idx}']
    if isinstance(formula, str) and formula.startswith("="):
         cell_value.value = formula
    else:
        cell_value.value = formula

    apply_data_style(cell_label)
    apply_data_style(cell_value, is_currency=True)

    if "TOTAL" in label or "BRUTA" in label or "SUBTOTAL" in label:
        cell_label.font = Font(bold=True)
        cell_value.font = Font(bold=True)
    row_idx += 1

# --- Ajuste final y guardado ---
# Autoajustar columnas en todas las hojas
for sheet in wb.worksheets:
    auto_adjust_columns(sheet)

# Guardar el archivo Excel
file_name = "Ejemplos_Valorizaciones.xlsx"
wb.save(file_name)

print(f"Archivo '{file_name}' generado exitosamente.")